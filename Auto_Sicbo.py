import more_functions.extra.Sicbo_Simulator as ss
import random
import openpyxl
import time

print('modules rendered')
tries = 200000
min_money = 500
max_money = 10000
print('loading workbook')
wb = openpyxl.load_workbook('sicbo results.xlsx')
print('loading worksheet')
ws = wb.active
print(ws, ws.max_row)
time.sleep(5)
result = []
print('starts gaming')
for t in range(tries):
    result.append(ss.sicbo(random.choice(['big', 'small']), random.choice(list(range(min_money, max_money)))))
    print(f'Game {t}: {result[-1]} (Playing section)')
    if t > tries - 1:
        break

wins = 0
money_back = 0
ws['F6'].value = 0
print('result making')
for a, b, c, d, e in result:
    profit = a - b
    money_back += profit
    add = ws.max_row + 1
    if profit >= 0:
        wins += 1
        ws[f'D{add}'].value = 'yes'
    else:
        ws[f'D{add}'].value = 'no'
    ws[f'A{add}'].value, ws[f'B{add}'].value, ws[f'C{add}'].value = [b, a, profit]
    ws['F6'].value += b
    print('Game', add - 1, ':', ws[f'A{add}'].value, ws[f'B{add}'].value, ws[f'C{add}'].value, '(Marking section)')

ws['E6'].value = len(result)
ws['G6'].value = money_back
ws['H6'].value = wins / len(result)
print(f"Winning rate: {100 * wins / len(result)}%")
if money_back >= 0:
    print(f"Money gain (per play): ${money_back / len(result)}")
else:
    print(f"Money lost (per play): ${-money_back / len(result)}")
wb.save('sicbo results.xlsx')
